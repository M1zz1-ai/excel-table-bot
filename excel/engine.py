"""Deterministic tabular engines for the Reformat and Compare wizards.

Philosophy (same as ``column_stats``): the LLM makes the *decisions* — which
source column maps to which template column, which key to join on, which columns
matter — but all row-by-row work (the fill, the diff, the sums) is deterministic
Python, never LLM-per-row. That keeps a 3k-row / 40-column file fast and exact.

Ported from the n8n "Excel Tables · TG Bot" workflow:
  * header-row detection (score = text-cells·2 + unique-cells over the first 20
    rows) — real spreadsheets have metadata/title rows above the header;
  * template headers: a header wrapped in ``*...*`` is MANDATORY; a ``CID`` column
    takes a constant from the template's first data row applied to every output row;
  * total/summary rows (Итого/Total/…) are excluded from the reformat fill.
The LLM-facing prompt/schema builders live here too; the actual model call is made
by the caller so this module stays pure and testable.
"""

from __future__ import annotations

import io
import logging
import re
from collections.abc import Callable
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rapidfuzz import fuzz, process

from .tools import _coerce_number

logger = logging.getLogger(__name__)

_TOTAL_RE = re.compile(
    r"^\s*(total|grand\s*total|subtotal|итого|итог|всего|всього|усього|разом|"
    r"підсумок|sum)\b",
    re.IGNORECASE,
)

REFORMAT_SYSTEM = (
    "You plan a spreadsheet reformat: decide which SOURCE column feeds each "
    "TEMPLATE column. You NEVER move data yourself — you only output the mapping "
    "plan as JSON matching the schema. Map only template columns; leave source_col "
    'and constant empty ("") when there is no good match.'
)
COMPARE_SYSTEM = (
    "You plan a comparison of two spreadsheets: pick the join key in each and the "
    "column pairs whose values should match. You output ONLY the plan JSON; the "
    "diff itself is computed by code."
)
TABLE_PLAN_SYSTEM = (
    "You read the RAW grid of a spreadsheet (a print/accounting export) and output "
    "a structure plan so code can extract the item rows deterministically. Real "
    "exports are messy: several title/metadata rows on top, the real header is not "
    "row 0, columns are sparse (values spread across non-adjacent columns), and a "
    "print layout often DUPLICATES the whole table horizontally to the right "
    "(a second block of the same columns). Rules:\n"
    "- Identify header_row (0-based index of the row holding the column captions).\n"
    "- Identify data_start_row (0-based index of the FIRST real item row).\n"
    "- Identify data_end_row (0-based index of the LAST real item row) — stop BEFORE "
    "any totals line ('Всього'/'Разом'/'Итого'/'Total'), payment footer, or notes.\n"
    "- List columns of the FIRST table block ONLY — one entry per real data column, "
    "with its 0-based column index and a short field name (prefer the caption, e.g. "
    "'Артикул', 'Товар', 'Кількість', 'Ціна', 'Сума'). IGNORE the duplicated block "
    "further right (higher column indices repeating the same captions).\n"
    "- Do NOT include empty spacer columns, totals, or metadata columns.\n"
    "Output ONLY the plan JSON."
)
COMPARE_ANSWER_SYSTEM = (
    "You are a data-diff analyst. Answer the user's question in 3-4 short sentences "
    "using ONLY the provided diff statistics and reconciliation — do not invent "
    "numbers. When the total sums differ, EXPLAIN the difference: state how the delta "
    "splits between products present in only one file and matched products whose sums "
    "differ, and name the largest contributors with their amount and reason "
    "(products, not codes). Talk money, concrete and concise."
)
LLM_PAIR_SYSTEM = (
    "You match rows of two spreadsheets that could not be joined automatically. "
    "You are given two lists of leftover keys (list A and list B). Return ONLY the "
    "pairs where a key from A and a key from B denote the SAME real-world item — "
    "including cross-language names (e.g. Russian 'гвоздь' == Ukrainian 'цвях'), "
    "heavy renames, and abbreviations. Never invent keys: every 'a' must appear "
    "verbatim in list A and every 'b' verbatim in list B. Each key may be used at "
    "most once. If unsure, leave a key out."
)

# Tier-2 fuzzy match threshold (rapidfuzz token_set_ratio, 0-100). Below this a
# pair is NOT auto-joined and falls through to the LLM tier / only-in buckets.
FUZZY_THRESHOLD = 85
# Fixed confidence assigned to an LLM-paired match (it is a judgement, not a score).
LLM_CONFIDENCE = 0.9
# Above this many unmatched keys on EITHER side, tier 3 is skipped (batch too big
# to pair reliably / cheaply in one call); recorded as ``llm_skipped`` in the diff.
LLM_MAX_KEYS = 200


# ---- tabular parsing (shared) ------------------------------------------


def sheet_rows(wb: Workbook) -> list[list[Any]]:
    """Read the active worksheet as a list of row value-lists."""
    return [list(row) for row in wb.active.iter_rows(values_only=True)]


def detect_header_row(rows: list[list[Any]]) -> int:
    """Pick the most header-like row index in the first 20 (n8n scoring)."""
    best_idx, best_score = 0, -1
    for i in range(min(20, len(rows))):
        vals = rows[i] or []
        text_count = sum(
            1 for v in vals if isinstance(v, str) and len(v.strip()) > 1 and _coerce_number(v) is None
        )
        unique = len({str(v).strip().lower() for v in vals if v is not None})
        score = text_count * 2 + unique
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _header_positions(rows: list[list[Any]], idx: int) -> list[tuple[int, str]]:
    """``(column_index, name)`` for each non-empty cell of header row ``idx``.

    Keeping the ORIGINAL column index is essential: a sparse header (gaps between
    labels) must still line data cells up under the right header. Compacting the
    names into a dense list — as an earlier version did — shifted every value
    under the wrong column for any header row with a gap.
    """
    if idx >= len(rows):
        return []
    return [
        (c, str(v).strip())
        for c, v in enumerate(rows[idx])
        if v is not None and str(v).strip() != ""
    ]


def extract_headers(rows: list[list[Any]], idx: int) -> list[str]:
    """Column names = the non-empty stringified cells of row ``idx`` (in order)."""
    return [name for _, name in _header_positions(rows, idx)]


def named_rows(rows: list[list[Any]], idx: int, headers: list[str]) -> list[dict[str, Any]]:
    """Rows below the header as ``{header: value}`` dicts (blank cells dropped).

    Each header is bound to its ORIGINAL column index (see :func:`_header_positions`),
    so sparse header rows no longer misalign the data.
    """
    positions = _header_positions(rows, idx)
    out: list[dict[str, Any]] = []
    for r in rows[idx + 1 :]:
        d: dict[str, Any] = {}
        for col, name in positions:
            v = r[col] if col < len(r) else None
            if v is None or str(v).strip() == "":
                continue
            d[name] = v
        if d:
            out.append(d)
    return out


def is_total_row(row: dict[str, Any]) -> bool:
    """True if any cell looks like a precomputed total/summary label."""
    return any(isinstance(v, str) and _TOTAL_RE.match(v) for v in row.values())


def load_table(wb: Workbook) -> tuple[list[str], list[dict[str, Any]]]:
    """Full pipeline: raw sheet -> (headers, named data rows, total rows dropped)."""
    rows = sheet_rows(wb)
    idx = detect_header_row(rows)
    headers = extract_headers(rows, idx)
    data = [r for r in named_rows(rows, idx, headers) if not is_total_row(r)]
    return headers, data


# ---- LLM-planned table extraction ---------------------------------------
#
# Heuristic header detection ("математически") breaks on real print/accounting
# exports — sparse columns, horizontal 2-up duplication, multi-row metadata. So
# the LLM LOOKS at the raw grid and returns a structure plan; extraction then
# follows the plan deterministically by column index. Falls back to the heuristic
# ``load_table`` path only if the model call fails or yields nothing usable.

TABLE_PLAN_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "header_row": {"type": "integer"},
        "data_start_row": {"type": "integer"},
        "data_end_row": {"type": "integer"},
        "columns": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {"col": {"type": "integer"}, "field": {"type": "string"}},
                "required": ["col", "field"],
                "additionalProperties": False,
            },
        },
        "notes": {"type": "string"},
    },
    "required": ["header_row", "data_start_row", "data_end_row", "columns", "notes"],
    "additionalProperties": False,
}


def _cellstr(value: Any, limit: int = 60) -> str:
    """Stringify a cell for the grid sample, truncated to keep the prompt small."""
    s = str(value).strip()
    return s if len(s) <= limit else s[: limit - 1] + "…"


def table_plan_prompt(rows: list[list[Any]], max_rows: int = 120, tail: int = 25) -> str:
    """Build the raw-grid sample prompt (original row + column indices preserved).

    Emits rows as ``{row, cells: {col_index: value}}`` maps so the model sees which
    column each value sits in (gaps and the horizontally-duplicated block included).
    The WHOLE grid is sent when it is at most ``max_rows`` tall — the model must see
    the end of the item block (and any totals/footer) to set ``data_end_row``. For
    a taller grid, the head plus the last ``tail`` rows are sent with the middle
    elided (row indices are absolute, so the plan still refers to real rows).
    """
    import json

    def cell_map(i: int, r: list[Any]) -> dict[str, Any]:
        return {
            "row": i,
            "cells": {
                str(c): _cellstr(v)
                for c, v in enumerate(r)
                if v is not None and str(v).strip() != ""
            },
        }

    n = len(rows)
    if n <= max_rows:
        grid = [cell_map(i, rows[i]) for i in range(n)]
        elision = ""
    else:
        head = max_rows - tail
        grid = [cell_map(i, rows[i]) for i in range(head)]
        grid += [cell_map(i, rows[i]) for i in range(n - tail, n)]
        elision = f" Rows {head}..{n - tail - 1} are elided (…); use the absolute 'row' indices."
    return (
        "Raw spreadsheet grid (cells keyed by 0-based COLUMN INDEX, empty cells "
        f"omitted).{elision}\n"
        f"{json.dumps(grid, ensure_ascii=False)}\n\n"
        "Return header_row, data_start_row, data_end_row (last item row, before any "
        "totals/footer), columns (first block only, one per real data column with "
        "its column index + short field name), and notes."
    )


def extract_by_plan(rows: list[list[Any]], plan: dict[str, Any]) -> tuple[list[str], list[dict[str, Any]]]:
    """Deterministically pull item rows out of the raw grid per a ``table_plan``.

    Reads by explicit column index from ``data_start_row`` down; drops rows where
    every planned column is empty, repeated header rows, total rows, and
    consecutive fully-identical rows (the vertical page-repeat artifact).
    """
    cols = [(int(c["col"]), str(c["field"]).strip()) for c in plan.get("columns", [])]
    headers = [name for _, name in cols]
    start = int(plan.get("data_start_row", 0))
    end_raw = plan.get("data_end_row")
    # data_end_row (inclusive) bounds the item block so trailing totals/footer rows
    # are excluded; ignore it if absent or nonsensical (<= start).
    end = int(end_raw) if isinstance(end_raw, (int, float)) and int(end_raw) >= start else len(rows) - 1
    header_labels = {name.lower() for _, name in cols if name}

    out: list[dict[str, Any]] = []
    prev: dict[str, Any] | None = None
    for r in rows[start : end + 1]:
        rec: dict[str, Any] = {}
        for col, name in cols:
            v = r[col] if col < len(r) else None
            if v is not None and str(v).strip() != "":
                rec[name] = v
        if not rec:
            continue
        values_lower = {str(v).strip().lower() for v in rec.values()}
        if values_lower and values_lower <= header_labels:  # a repeated header row
            continue
        if is_total_row(rec):
            continue
        if rec == prev:  # vertical page-repeat duplicate
            continue
        out.append(rec)
        prev = rec
    return headers, out


def load_table_planned(
    wb: Workbook,
    *,
    plan_fn: Callable[[list[list[Any]]], dict[str, Any]] | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Load a data table via the LLM ``table_plan`` step, or fall back to heuristics.

    ``plan_fn`` (injected — the module owns no model call) takes the raw grid rows
    and returns a plan matching :data:`TABLE_PLAN_SCHEMA`. On any failure, or if
    the plan yields no rows, this falls back to the heuristic :func:`load_table`.
    """
    rows = sheet_rows(wb)
    if plan_fn is not None:
        try:
            plan = plan_fn(rows)
            headers, data = extract_by_plan(rows, plan)
            if data:
                return headers, data
            logger.warning("table_plan produced no rows; falling back to heuristic loader")
        except Exception:  # LLM/parse failure must not break ingestion
            logger.warning("table_plan step failed; falling back to heuristic loader", exc_info=True)
    return load_table(wb)


# ---- reformat engine ----------------------------------------------------


def _template_from_rows(rows: list[list[Any]], header_idx: int, data_idx: int) -> dict[str, Any]:
    """Template schema from a specific header row: names, ``*mandatory*``, CID const.

    ``*name*`` marks a mandatory column; a ``CID`` column takes the constant from the
    template's first data row (``data_idx``) and applies it to every output row.
    """
    header_row = rows[header_idx] if 0 <= header_idx < len(rows) else []
    headers: list[dict[str, Any]] = []
    cid_constant: str | None = None
    for pos, raw in enumerate(header_row):
        if raw is None or str(raw).strip() == "":
            continue
        s = str(raw).strip()
        mandatory = bool(re.fullmatch(r"\*.+\*", s))
        name = s[1:-1].strip() if mandatory else s
        headers.append({"name": name, "mandatory": mandatory})
        if name.upper() == "CID" and 0 <= data_idx < len(rows):
            cell = rows[data_idx][pos] if pos < len(rows[data_idx]) else None
            if cell is not None and str(cell).strip():
                cid_constant = str(cell).strip()
    return {"headers": headers, "cid_constant": cid_constant}


def parse_template(wb: Workbook) -> dict[str, Any]:
    """Template schema from the FIRST row (heuristic fallback path)."""
    return _template_from_rows(sheet_rows(wb), 0, 1)


def parse_template_planned(
    wb: Workbook,
    *,
    plan_fn: Callable[[list[list[Any]]], dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Template schema via the LLM ``table_plan`` (real templates have metadata rows).

    ``plan_fn`` locates the header row (and first data row) in a messy template the
    same way compare/data ingestion does; ``*mandatory*`` and CID semantics are then
    applied on top of that planned header row. Falls back to row 0 on any failure.
    """
    rows = sheet_rows(wb)
    if plan_fn is not None:
        try:
            plan = plan_fn(rows)
            header_idx = int(plan.get("header_row", 0))
            data_idx = int(plan.get("data_start_row", header_idx + 1))
            tpl = _template_from_rows(rows, header_idx, data_idx)
            if tpl["headers"]:
                return tpl
            logger.warning("template plan produced no headers; falling back to row 0")
        except Exception:  # LLM/parse failure must not break the wizard
            logger.warning("template plan step failed; falling back to row 0", exc_info=True)
    return _template_from_rows(rows, 0, 1)


REFORMAT_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "columns": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "template_col": {"type": "string"},
                    "source_col": {"type": "string"},
                    "constant": {"type": "string"},
                    "fuzzy": {"type": "boolean"},
                },
                "required": ["template_col", "source_col", "constant", "fuzzy"],
                "additionalProperties": False,
            },
        },
        "key_column": {"type": "string"},
    },
    "required": ["columns", "key_column"],
    "additionalProperties": False,
}


def reformat_prompt(
    template: dict[str, Any], source_cols: list[str], sample: list[dict[str, Any]], instruction: str
) -> str:
    """Build the mapping-plan prompt (LLM decides the column map; code fills rows)."""
    import json

    tpl = [h["name"] for h in template["headers"]]
    mandatory = [h["name"] for h in template["headers"] if h["mandatory"]]
    return (
        "Map SOURCE columns onto TEMPLATE columns for a spreadsheet reformat.\n\n"
        f"TEMPLATE columns:\n{json.dumps(tpl, ensure_ascii=False)}\n\n"
        f"Mandatory template columns (must be filled if at all possible):\n"
        f"{json.dumps(mandatory, ensure_ascii=False)}\n\n"
        f"SOURCE columns:\n{json.dumps(source_cols, ensure_ascii=False)}\n\n"
        f"SOURCE sample rows:\n{json.dumps(sample[:15], ensure_ascii=False)}\n\n"
        f"User instruction:\n{instruction}\n\n"
        "For each template column output one entry: template_col (copy the template "
        "column name EXACTLY as given above, with no extra words), source_col (the "
        'exact source column name to copy, or "" if none), constant (a fixed value '
        'to use instead, or ""), fuzzy (true if the source match is approximate). '
        'Set key_column to the template column that identifies a row (or ""). Only '
        "map columns that exist in the template."
    )


def _clean_tpl_name(value: Any) -> str:
    """Strip a trailing ``(mandatory)`` decoration the model may echo into a name."""
    return re.sub(r"\s*\(mandatory\)\s*$", "", str(value), flags=re.IGNORECASE).strip()


def apply_mapping(
    template: dict[str, Any], source_rows: list[dict[str, Any]], plan: dict[str, Any]
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Fill the template's columns from every source row per the mapping plan.

    Deterministic (no LLM). Returns ``(rows, notes)`` where notes carry the
    fuzzy-matched fields, mandatory template fields with no mapping, and the
    duplicate-key count.
    """
    headers = template["headers"]
    cid = template.get("cid_constant")
    # Match on the clean template name: models sometimes echo a decorated column
    # (e.g. "Артикул (mandatory)"), which must still map to "Артикул".
    by_tpl = {_clean_tpl_name(c["template_col"]): c for c in plan.get("columns", [])}
    key_column = plan.get("key_column") or ""

    out: list[dict[str, Any]] = []
    key_counts: dict[str, int] = {}
    for src in source_rows:
        row: dict[str, Any] = {}
        for h in headers:
            name = h["name"]
            spec = by_tpl.get(name, {})
            if name.upper() == "CID" and cid is not None:
                row[name] = cid
            elif spec.get("constant"):
                row[name] = spec["constant"]
            elif spec.get("source_col"):
                row[name] = src.get(spec["source_col"])
            else:
                row[name] = None
        out.append(row)
        if key_column and row.get(key_column) not in (None, ""):
            k = str(row[key_column]).strip()
            key_counts[k] = key_counts.get(k, 0) + 1

    names = [h["name"] for h in headers]
    mapped_cols: set[str] = set()
    for h in headers:
        name = h["name"]
        spec = by_tpl.get(name, {})
        if (name.upper() == "CID" and cid is not None) or spec.get("constant") or spec.get("source_col"):
            mapped_cols.add(name)

    fill_counts = {n: 0 for n in names}
    for row in out:
        for n in names:
            if row.get(n) not in (None, ""):
                fill_counts[n] += 1
    total_cells = len(out) * len(names)
    overall_fill = round(sum(fill_counts.values()) / total_cells, 4) if total_cells else 0.0

    mapped = {name for name, s in by_tpl.items() if s.get("source_col") or s.get("constant")}
    notes = {
        "fuzzy_fields": [_clean_tpl_name(c["template_col"]) for c in plan.get("columns", []) if c.get("fuzzy")],
        "unmapped_mandatory": [
            h["name"]
            for h in headers
            if h["mandatory"] and h["name"] not in mapped and h["name"].upper() != "CID"
        ],
        "duplicate_keys": sum(1 for c in key_counts.values() if c > 1),
        "row_count": len(out),
        # fill diagnostics — used to refuse silent empty output.
        "template_cols": len(names),
        "mapped_columns": len(mapped_cols),
        "unmapped_columns": [n for n in names if n not in mapped_cols],
        "empty_columns": [n for n, c in fill_counts.items() if c == 0],
        "column_fill": fill_counts,
        "overall_fill_pct": overall_fill,
        "mapping_pairs": [
            (_clean_tpl_name(c["template_col"]),
             c.get("source_col") or (f"const:{c['constant']}" if c.get("constant") else "—"))
            for c in plan.get("columns", [])
        ],
    }
    return out, notes


def reformat_quality(notes: dict[str, Any]) -> str:
    """Grade a reformat result so the bot never sends a silently-empty file.

    ``"empty"`` — nothing mapped / no rows / all cells blank (send NO file);
    ``"weak"``  — more than half the template columns got no mapping (send file +
    explanation); ``"ok"`` — normal success.
    """
    if notes["row_count"] == 0 or notes.get("mapped_columns", 0) == 0 or notes.get("overall_fill_pct", 0) == 0:
        return "empty"
    total = notes.get("template_cols", 0)
    unmapped = len(notes.get("unmapped_columns", []))
    if total and unmapped / total > 0.5:
        return "weak"
    return "ok"


def build_reformat_xlsx(template: dict[str, Any], rows: list[dict[str, Any]]) -> bytes:
    """Write the reformatted rows into a fresh .xlsx (template header order)."""
    names = [h["name"] for h in template["headers"]]
    wb = Workbook()
    ws = wb.active
    ws.append(names)
    for r in rows:
        ws.append([r.get(n) for n in names])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- compare engine -----------------------------------------------------


COMPARE_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "key_a": {"type": "string"},
        "key_b": {"type": "string"},
        "compare_columns": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "a": {"type": "string"},
                    "b": {"type": "string"},
                    "label": {"type": "string"},
                },
                "required": ["a", "b", "label"],
                "additionalProperties": False,
            },
        },
        "sum_column_a": {"type": "string"},
        "sum_column_b": {"type": "string"},
    },
    "required": ["key_a", "key_b", "compare_columns", "sum_column_a", "sum_column_b"],
    "additionalProperties": False,
}


def compare_prompt(
    headers_a: list[str],
    headers_b: list[str],
    sample_a: list[dict[str, Any]],
    sample_b: list[dict[str, Any]],
    question: str,
) -> str:
    """Build the compare-plan prompt (LLM picks join key + columns; code diffs)."""
    import json

    return (
        "Plan a comparison of two spreadsheets (A and B).\n\n"
        f"A columns: {json.dumps(headers_a, ensure_ascii=False)}\n"
        f"A sample: {json.dumps(sample_a[:10], ensure_ascii=False)}\n\n"
        f"B columns: {json.dumps(headers_b, ensure_ascii=False)}\n"
        f"B sample: {json.dumps(sample_b[:10], ensure_ascii=False)}\n\n"
        f"User question: {question}\n\n"
        "Choose key_a / key_b — the column in each file that identifies the SAME "
        "item. STRONGLY PREFER a stable identifier column (артикул / код / code / "
        "SKU / ID / barcode: short alphanumeric values, one per row, highly unique) "
        "over descriptive name columns. The two files may be in different languages "
        "(e.g. Russian vs Ukrainian), so product NAMES will not match literally "
        "while codes will — join on the code. Only fall back to a name column if no "
        "code-like column exists in BOTH files. Then choose compare_columns (pairs "
        "of a/b columns whose values should match, each with a short label) and "
        "sum_column_a / sum_column_b (the numeric column whose totals the user cares "
        'about, or "").'
    )


# ---- tiered key matcher -------------------------------------------------
#
# Real files rarely share exact string keys: the same product is named slightly
# differently, has typos/abbreviations, or is even in another language (RU vs UA).
# Exact-join then dumps everything into only-in-A / only-in-B and the report looks
# broken. So the join runs in three tiers, cheapest first:
#   1. exact on a NORMALIZED key (case/space/punct/ё/confusable folding);
#   2. fuzzy (rapidfuzz) with a global greedy one-to-one assignment;
#   3. LLM residual pairing (injected callable) for translations / heavy renames.

# Cyrillic letters that look identical to a latin one, folded to latin so a key
# typed with mixed alphabets ("Аpple" with a Cyrillic А) matches its twin. The
# mapping is injective, so distinct real words never collapse into one another.
_CONFUSABLES = str.maketrans(
    {
        "а": "a", "е": "e", "о": "o", "р": "p", "с": "c", "у": "y", "х": "x",
        "к": "k", "в": "b", "н": "h", "м": "m", "т": "t", "і": "i", "ѕ": "s",
    }
)
# Standalone measurement tokens dropped as noise (kept conservative — only tokens
# that are unambiguously units, never single ambiguous letters like "г"/"м"/"л").
_UNIT_TOKENS = {
    "кг", "kg", "гр", "мг", "mg", "мл", "ml", "шт", "pcs", "pc", "уп",
    "мм", "mm", "см", "cm", "м2", "м3",
}
_PUNCT_RE = re.compile(r"[^\w\s]", re.UNICODE)


def normalize_key(value: Any) -> str:
    """Fold a join key to a canonical form for exact/fuzzy matching.

    Lowercase, ``ё``->``е``, unify latin/cyrillic look-alike letters, strip
    punctuation, drop standalone unit tokens, and collapse whitespace. Two keys
    that a human would call "the same" (bar language) should normalize equal.
    """
    s = str(value).strip().lower().replace("ё", "е")
    s = _PUNCT_RE.sub(" ", s).replace("_", " ")
    # Drop unit tokens in their natural spelling BEFORE latin/cyrillic folding
    # (folding would otherwise turn "кг" into "kг" and miss the set).
    tokens = [t.translate(_CONFUSABLES) for t in s.split() if t not in _UNIT_TOKENS]
    return " ".join(tokens)


def match_keys(
    keys_a: list[str],
    keys_b: list[str],
    *,
    llm_pair: Callable[[list[str], list[str]], list[dict[str, str]]] | None = None,
) -> dict[str, Any]:
    """Join two key lists with the exact/fuzzy/LLM tiers; return the assignment.

    ``llm_pair`` (tier 3, optional) is an injected callable — the module never
    owns a model call. It takes the still-unmatched ``(keys_a, keys_b)`` and
    returns ``[{"a": .., "b": ..}]`` pairs; hallucinated keys and one-to-many
    reuse are validated out here. Returns a dict with ``matched`` (each carrying
    ``method`` + ``confidence``), ``unmatched_a`` / ``unmatched_b``,
    ``match_counts`` and ``llm_skipped``.
    """
    a = list(dict.fromkeys(k for k in keys_a if str(k).strip() != ""))
    b = list(dict.fromkeys(k for k in keys_b if str(k).strip() != ""))
    norm_a = {k: normalize_key(k) for k in a}
    norm_b = {k: normalize_key(k) for k in b}

    matched: list[dict[str, Any]] = []
    used_a: set[str] = set()
    used_b: set[str] = set()

    # Tier 1 — exact on normalized keys (first B wins for a normalized value).
    b_by_norm: dict[str, str] = {}
    for k in b:
        b_by_norm.setdefault(norm_b[k], k)
    for k in a:
        nk = norm_a[k]
        cand = b_by_norm.get(nk)
        if nk and cand is not None and cand not in used_b:
            matched.append({"key_a": k, "key_b": cand, "method": "exact", "confidence": 1.0})
            used_a.add(k)
            used_b.add(cand)

    # Tier 2 — fuzzy, global greedy one-to-one above FUZZY_THRESHOLD.
    rem_a = [k for k in a if k not in used_a]
    rem_b = [k for k in b if k not in used_b]
    if rem_a and rem_b:
        choices = {k: norm_b[k] for k in rem_b if norm_b[k]}
        candidates: list[tuple[float, str, str]] = []
        for ka in rem_a:
            if not norm_a[ka]:
                continue
            for _s, score, kb in process.extract(
                norm_a[ka], choices, scorer=fuzz.token_set_ratio,
                score_cutoff=FUZZY_THRESHOLD, limit=5,
            ):
                candidates.append((score, ka, kb))
        candidates.sort(key=lambda c: c[0], reverse=True)
        for score, ka, kb in candidates:
            if ka in used_a or kb in used_b:
                continue
            matched.append(
                {"key_a": ka, "key_b": kb, "method": "fuzzy", "confidence": round(score / 100, 4)}
            )
            used_a.add(ka)
            used_b.add(kb)

    # Tier 3 — LLM residual pairing (translations / heavy renames).
    rem_a = [k for k in a if k not in used_a]
    rem_b = [k for k in b if k not in used_b]
    llm_skipped = False
    if rem_a and rem_b and llm_pair is not None:
        if len(rem_a) > LLM_MAX_KEYS or len(rem_b) > LLM_MAX_KEYS:
            llm_skipped = True
        else:
            set_a, set_b = set(rem_a), set(rem_b)
            for pair in llm_pair(list(rem_a), list(rem_b)) or []:
                ka, kb = pair.get("a"), pair.get("b")
                if ka in set_a and kb in set_b and ka not in used_a and kb not in used_b:
                    matched.append(
                        {"key_a": ka, "key_b": kb, "method": "llm", "confidence": LLM_CONFIDENCE}
                    )
                    used_a.add(ka)
                    used_b.add(kb)

    counts = {
        "exact": sum(1 for m in matched if m["method"] == "exact"),
        "fuzzy": sum(1 for m in matched if m["method"] == "fuzzy"),
        "llm": sum(1 for m in matched if m["method"] == "llm"),
    }
    return {
        "matched": matched,
        "unmatched_a": [k for k in a if k not in used_a],
        "unmatched_b": [k for k in b if k not in used_b],
        "match_counts": counts,
        "llm_skipped": llm_skipped,
    }


LLM_PAIR_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "pairs": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {"a": {"type": "string"}, "b": {"type": "string"}},
                "required": ["a", "b"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["pairs"],
    "additionalProperties": False,
}


def llm_pair_prompt(keys_a: list[str], keys_b: list[str]) -> str:
    """Build the tier-3 residual-pairing prompt (LLM returns same-item key pairs)."""
    import json

    return (
        "Two spreadsheets left these keys unmatched after exact + fuzzy joining.\n\n"
        f"List A ({len(keys_a)}): {json.dumps(keys_a, ensure_ascii=False)}\n\n"
        f"List B ({len(keys_b)}): {json.dumps(keys_b, ensure_ascii=False)}\n\n"
        "Return pairs {a, b} where the A key and the B key are the SAME real-world "
        "item (translations, renames, abbreviations). Use each key at most once; "
        "copy keys verbatim; omit anything you are unsure about."
    )


def _index_by_key(rows: list[dict[str, Any]], key: str) -> dict[str, dict[str, Any]]:
    idx: dict[str, dict[str, Any]] = {}
    for r in rows:
        k = r.get(key)
        if k is None or str(k).strip() == "":
            continue
        idx[str(k).strip()] = r
    return idx


# A join key must actually join: the LLM plan sometimes picks a descriptive name
# column (which fails across languages, RU vs UA) over the stable code column. If
# the planned key barely overlaps, deterministically prefer an id-like column pair
# (high per-row uniqueness) with strong normalized-value overlap.
KEY_OVERLAP_MIN = 0.5      # planned key is kept if it already joins this fraction
KEY_MIN_UNIQUENESS = 0.8  # a candidate key column must be near one-value-per-row
KEY_OVERRIDE_ABS = 0.7    # and overlap at least this much to override the plan
KEY_OVERRIDE_MARGIN = 0.1


def _column_names(rows: list[dict[str, Any]]) -> list[str]:
    seen: dict[str, None] = {}
    for r in rows:
        for k in r:
            seen.setdefault(k, None)
    return list(seen)


def _norm_list(rows: list[dict[str, Any]], col: str) -> list[str]:
    return [
        normalize_key(r[col])
        for r in rows
        if r.get(col) is not None and str(r.get(col)).strip() != ""
    ]


def _uniqueness(values: list[str]) -> float:
    return len(set(values)) / len(values) if values else 0.0


def _looks_ordinal(rows: list[dict[str, Any]], col: str) -> bool:
    """True if a column is just a row counter (1,2,3,…) — a positional artifact.

    Such a column overlaps perfectly across two files but joins by ROW POSITION,
    not identity, so it must never be chosen as a join key.
    """
    nums: list[int] = []
    for r in rows:
        v = r.get(col)
        if v is None or str(v).strip() == "":
            continue
        n = _coerce_number(v)
        if n is None or n != int(n):
            return False  # a non-integer value -> not a plain counter
        nums.append(int(n))
    if len(nums) < 3:
        return False
    s = sorted(nums)
    return s[0] in (0, 1) and s == list(range(s[0], s[0] + len(s)))


def resolve_join_keys(
    rows_a: list[dict[str, Any]], rows_b: list[dict[str, Any]], plan: dict[str, Any]
) -> tuple[str, str]:
    """Return the join keys to use — the plan's, unless a better id-like pair exists.

    Deterministic safety net for the language-mismatch case: if the planned key
    joins less than :data:`KEY_OVERLAP_MIN` of rows, scan for a column pair that is
    highly unique on both sides (an article/code/SKU column) and whose normalized
    values overlap strongly, and use that instead.
    """
    ka, kb = plan.get("key_a"), plan.get("key_b")
    cols_a, cols_b = _column_names(rows_a), _column_names(rows_b)
    lists_a = {c: _norm_list(rows_a, c) for c in cols_a}
    lists_b = {c: _norm_list(rows_b, c) for c in cols_b}
    sets_a = {c: set(v) for c, v in lists_a.items()}
    sets_b = {c: set(v) for c, v in lists_b.items()}

    def overlap(ca: str, cb: str) -> float:
        sa, sb = sets_a.get(ca, set()), sets_b.get(cb, set())
        return len(sa & sb) / min(len(sa), len(sb)) if sa and sb else 0.0

    base = overlap(ka, kb) if ka in sets_a and kb in sets_b else 0.0
    if base >= KEY_OVERLAP_MIN:
        return ka, kb  # planned key already joins well
    # Value columns being compared/summed are not join keys — exclude them so a
    # coincidentally-overlapping amount column is never chosen as the key.
    comp = plan.get("compare_columns", [])
    excl_a = {plan.get("sum_column_a")} | {c.get("a") for c in comp}
    excl_b = {plan.get("sum_column_b")} | {c.get("b") for c in comp}
    best_a, best_b, best = ka or "", kb or "", base
    for ca in cols_a:
        if ca in excl_a or _uniqueness(lists_a[ca]) < KEY_MIN_UNIQUENESS:
            continue
        if _looks_ordinal(rows_a, ca):  # a row-number column, not an identity
            continue
        for cb in cols_b:
            if cb in excl_b or _uniqueness(lists_b[cb]) < KEY_MIN_UNIQUENESS:
                continue
            if _looks_ordinal(rows_b, cb):
                continue
            ov = overlap(ca, cb)
            if ov >= KEY_OVERRIDE_ABS and ov > best + KEY_OVERRIDE_MARGIN:
                best_a, best_b, best = ca, cb, ov
    return best_a, best_b


def _descriptive_field(rows: list[dict[str, Any]], exclude: str | None) -> str | None:
    """The most descriptive column (longest total string length) — the product name."""
    scores: dict[str, int] = {}
    for r in rows:
        for k, v in r.items():
            if k == exclude or not isinstance(v, str):
                continue
            scores[k] = scores.get(k, 0) + len(v.strip())
    return max(scores, key=lambda k: scores[k]) if scores else None


def _row_name(row: dict[str, Any], field: str | None, fallback: str) -> str:
    if field and row.get(field) not in (None, ""):
        return str(row[field])
    return fallback


def reconcile_sum(diff: dict[str, Any], plan: dict[str, Any]) -> list[dict[str, Any]]:
    """Decompose the sum delta (ΣA − ΣB) into per-product money contributions.

    Deterministic, no LLM. Answers "which product made the totals differ, and why":
      * a matched product whose sums differ contributes ``sumA − sumB`` (reason = the
        underlying compare-field diffs, e.g. «Кол-во 24 → 20», else «сумма отличается»);
      * an only-in-A product contributes ``+sumA`` (reason «есть только в A»);
      * an only-in-B product contributes ``−sumB`` (reason «есть только в B»).
    A residual line «прочее / округление» absorbs any rounding gap so the
    contributions always total Δ. Sorted by absolute impact, non-zero only.
    """
    sca, scb = plan.get("sum_column_a") or "", plan.get("sum_column_b") or ""
    comp = plan.get("compare_columns", [])
    sum_labels = {c["label"] for c in comp if c.get("a") == sca or c.get("b") == scb}

    mm_by_key: dict[str, list[dict[str, Any]]] = {}
    for m in diff["mismatches"]:
        mm_by_key.setdefault(m["key_a"], []).append(m)

    contribs: list[dict[str, Any]] = []
    for m in diff["matched"]:
        c = round((m.get("sum_a") or 0.0) - (m.get("sum_b") or 0.0), 2)
        if c == 0:
            continue
        causes = [
            f"{x['field']} {x['value_a']} → {x['value_b']}"
            for x in mm_by_key.get(m["key_a"], [])
            if x["field"] not in sum_labels
        ]
        contribs.append(
            {
                "name": m["name"],
                "code": m["key_a"],
                "contribution": c,
                "reason": "; ".join(causes) if causes else "сумма отличается",
            }
        )
    for it in diff["only_in_a"]:
        v = _coerce_number(it.get(sca))
        if v:
            contribs.append(
                {
                    "name": _row_name(it, diff.get("name_field_a"), str(it.get("key", ""))),
                    "code": str(it.get("key", "")),
                    "contribution": round(v, 2),
                    "reason": "есть только в A",
                }
            )
    for it in diff["only_in_b"]:
        v = _coerce_number(it.get(scb))
        if v:
            contribs.append(
                {
                    "name": _row_name(it, diff.get("name_field_b"), str(it.get("key", ""))),
                    "code": str(it.get("key", "")),
                    "contribution": round(-v, 2),
                    "reason": "есть только в B",
                }
            )

    delta = diff.get("delta")
    if delta is not None:
        residual = round(delta - sum(c["contribution"] for c in contribs), 2)
        if abs(residual) >= 0.01:
            contribs.append(
                {"name": "прочее / округление", "code": "", "contribution": residual, "reason": "остаток"}
            )
    contribs.sort(key=lambda c: abs(c["contribution"]), reverse=True)
    return contribs


def compare_discrepancy_block(diff: dict[str, Any], top: int = 5) -> str:
    """Deterministic plain-text top-discrepancy list for the Telegram message.

    Product-centric, real numbers (from the reconciliation, not the LLM). Capped at
    ``top`` products by money impact so the message fits Telegram; the rest live in
    the file's «Расхождения» sheet.
    """
    recon = [c for c in diff.get("reconciliation", []) if c["code"]]  # drop residual line
    if not recon:
        return ""
    shown = recon[:top]
    lines = [f"Товары с расхождениями (топ-{len(shown)} по деньгам):"]
    for c in shown:
        lines.append(f"• {c['name']} ({c['code']}): {c['reason']} ({c['contribution']:+.2f})")
    n_products = len({m["key_a"] for m in diff["mismatches"]})
    lines.append(
        f'Всего расхождений: {len(diff["mismatches"])} у {n_products} товаров. '
        'Полный список — в файле, лист "Расхождения".'
    )
    return "\n".join(lines)


def reconcile_line(diff: dict[str, Any], top: int = 10) -> str:
    """Compact Russian reconciliation fed to the answer LLM (real numbers only)."""
    recon = diff.get("reconciliation", [])
    if not recon:
        return ""
    plus = round(sum(c["contribution"] for c in recon if c["contribution"] > 0), 2)
    minus = round(sum(c["contribution"] for c in recon if c["contribution"] < 0), 2)
    lines = [f"Разница {diff.get('delta')} = вклады товаров (плюсы +{plus}, минусы {minus}):"]
    for c in recon[:top]:
        lines.append(f"{c['name']} [{c['code']}]: {c['contribution']:+.2f} — {c['reason']}")
    return "\n".join(lines)


def diff_tables(
    rows_a: list[dict[str, Any]],
    rows_b: list[dict[str, Any]],
    plan: dict[str, Any],
    *,
    llm_pair: Callable[[list[str], list[str]], list[dict[str, str]]] | None = None,
) -> dict[str, Any]:
    """Deterministic diff over TIER-MATCHED key pairs (exact/fuzzy/LLM).

    Keys no longer have to be byte-identical: :func:`match_keys` joins A<->B via
    the three tiers, and the per-field diff then runs on the matched pairs exactly
    as before (numeric coercion via ``_coerce_number``). ``llm_pair`` is the
    injected tier-3 callable (the caller wires the real model); ``None`` disables
    tier 3. Each mismatch carries ``key_a``/``key_b``/``match_method``/``confidence``.
    """
    key_a, key_b = resolve_join_keys(rows_a, rows_b, plan)
    ia = _index_by_key(rows_a, key_a)
    ib = _index_by_key(rows_b, key_b)
    name_field_a = _descriptive_field(rows_a, key_a)
    name_field_b = _descriptive_field(rows_b, key_b)
    sum_col_a = plan.get("sum_column_a") or ""
    sum_col_b = plan.get("sum_column_b") or ""
    # Guarantee the sum column pair is compared (the money delta must always show up
    # in «Расхождения»), regardless of whether the LLM plan included it.
    compare_cols = list(plan.get("compare_columns", []))
    if sum_col_a and sum_col_b and not any(
        c.get("a") == sum_col_a and c.get("b") == sum_col_b for c in compare_cols
    ):
        compare_cols.append({"a": sum_col_a, "b": sum_col_b, "label": "Сумма"})
    plan = {**plan, "compare_columns": compare_cols}
    match = match_keys(list(ia), list(ib), llm_pair=llm_pair)
    used_a = {m["key_a"] for m in match["matched"]}
    used_b = {m["key_b"] for m in match["matched"]}

    only_in_a = [{"key": k, **ia[k]} for k in sorted(set(ia) - used_a)]
    only_in_b = [{"key": k, **ib[k]} for k in sorted(set(ib) - used_b)]

    mismatches: list[dict[str, Any]] = []
    for m in match["matched"]:
        ra, rb = ia[m["key_a"]], ib[m["key_b"]]
        m["name"] = _row_name(ra, name_field_a, m["key_a"])  # product-centric report
        m["sum_a"] = _coerce_number(ra.get(sum_col_a)) if sum_col_a else None
        m["sum_b"] = _coerce_number(rb.get(sum_col_b)) if sum_col_b else None
        for col in plan.get("compare_columns", []):
            va, vb = ra.get(col["a"]), rb.get(col["b"])
            na, nb = _coerce_number(va), _coerce_number(vb)
            delta: float | None
            if na is not None and nb is not None:
                delta = round(na - nb, 10)
                if delta == 0:
                    continue
            elif str(va or "").strip() == str(vb or "").strip():
                continue
            else:
                delta = None
            mismatches.append(
                {
                    "name": m["name"],
                    "key_a": m["key_a"],
                    "key_b": m["key_b"],
                    "field": col["label"],
                    "value_a": va,
                    "value_b": vb,
                    "delta": delta,
                    "match_method": m["method"],
                    "confidence": m["confidence"],
                }
            )

    sum_a = _sum_col(rows_a, sum_col_a)
    sum_b = _sum_col(rows_b, sum_col_b)
    diff = {
        "count_a": len(rows_a),
        "count_b": len(rows_b),
        "matched": match["matched"],
        "match_counts": match["match_counts"],
        "llm_skipped": match["llm_skipped"],
        "key_field_a": key_a,
        "key_field_b": key_b,
        "name_field_a": name_field_a,
        "name_field_b": name_field_b,
        "sum_field_a": sum_col_a,
        "sum_field_b": sum_col_b,
        "only_in_a": only_in_a,
        "only_in_b": only_in_b,
        "mismatches": mismatches,
        "sum_a": sum_a,
        "sum_b": sum_b,
        "delta": round(sum_a - sum_b, 10) if sum_a is not None and sum_b is not None else None,
    }
    diff["reconciliation"] = reconcile_sum(diff, plan)
    return diff


def _sum_col(rows: list[dict[str, Any]], col: str) -> float | None:
    if not col:
        return None
    nums = [n for r in rows if (n := _coerce_number(r.get(col))) is not None]
    return round(sum(nums), 10) if nums else 0.0


_METHOD_RU = {"exact": "точно", "fuzzy": "похоже", "llm": "по смыслу"}


def _finish_sheet(ws: Worksheet, widths: list[float]) -> None:
    """Bold + freeze the header row and set column widths so nothing truncates."""
    ws.freeze_panes = "A2"
    for c in ws[1]:
        c.font = Font(bold=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _ordered_fields(items: list[dict[str, Any]], name_field: str | None) -> list[str]:
    """Field order for a product listing: the name column first, then the rest.

    The synthetic ``key`` column (a duplicate of the join key) is dropped so the
    sheet shows real fields (артикул / кол-во / цена / сумма), not bare codes.
    """
    seen: dict[str, None] = {}
    for it in items:
        for k in it:
            if k != "key":
                seen.setdefault(k, None)
    fields = list(seen)
    if name_field and name_field in fields:
        fields.remove(name_field)
        fields = [name_field] + fields
    return fields


def build_compare_xlsx(diff: dict[str, Any]) -> bytes:
    """Human, product-centric report in Russian.

    Sheet order (the user opens the file for the per-product diff, so it comes first
    and is the active sheet): «Расхождения» / «Почему разница» / «Итог» /
    «Совпавшие (неточно)» (audit, when any) / «Только в A» / «Только в B» (when
    non-empty). Every sheet has a bold frozen header and non-truncating widths.
    """
    mc = diff["match_counts"]
    recon = diff.get("reconciliation", [])
    impact = {c["code"]: abs(c["contribution"]) for c in recon}

    wb = Workbook()
    # «Расхождения» first = the active sheet on open. Ordered by money impact.
    расх = wb.active
    расх.title = "Расхождения"
    расх.append(["Товар", "Код", "Поле", "Значение A", "Значение B", "Разница"])
    for m in sorted(diff["mismatches"], key=lambda m: impact.get(m.get("key_a"), 0), reverse=True):
        расх.append(
            [m.get("name"), m.get("key_a"), m["field"], m["value_a"], m["value_b"], m["delta"]]
        )
    _finish_sheet(расх, [50, 16, 16, 18, 18, 12])

    # «Почему разница» — Δ decomposed into per-product money contributions.
    почему = wb.create_sheet("Почему разница")
    почему.append(["Товар", "Артикул", "Вклад в разницу", "Причина"])
    for c in recon:
        почему.append([c["name"], c["code"], c["contribution"], c["reason"]])
    total_row = почему.max_row + 1
    почему.append(["ИТОГО разница", "", round(sum(c["contribution"] for c in recon), 2), ""])
    for cell in почему[total_row]:
        cell.font = Font(bold=True)
    _finish_sheet(почему, [50, 16, 18, 40])

    итог = wb.create_sheet("Итог")
    итог.append(["Показатель", "Значение"])
    for label, value in (
        ("Строк в A", diff["count_a"]),
        ("Строк в B", diff["count_b"]),
        ("Совпало (всего)", len(diff["matched"])),
        ("  точно", mc["exact"]),
        ("  похоже", mc["fuzzy"]),
        ("  по смыслу", mc["llm"]),
        ("Расхождений по полям", len(diff["mismatches"])),
        ("Только в A", len(diff["only_in_a"])),
        ("Только в B", len(diff["only_in_b"])),
        ("Сумма A", diff["sum_a"]),
        ("Сумма B", diff["sum_b"]),
        ("Разница", diff["delta"]),
    ):
        итог.append([label, value])
    if diff.get("llm_skipped"):
        итог.append(["Сопоставление по смыслу", "пропущено (слишком много строк)"])
    итог.append(['Детали по товарам — на листах "Расхождения" и "Почему разница"', ""])
    _finish_sheet(итог, [52, 18])

    nonexact = [m for m in diff["matched"] if m["method"] != "exact"]
    if nonexact:
        sov = wb.create_sheet("Совпавшие (неточно)")
        sov.append(["Товар", "Код A", "Код B", "Метод", "Уверенность"])
        for m in nonexact:
            sov.append(
                [m.get("name"), m["key_a"], m["key_b"], _METHOD_RU.get(m["method"], m["method"]),
                 m["confidence"]]
            )
        _finish_sheet(sov, [50, 16, 16, 12, 12])

    for title, items, name_field in (
        ("Только в A", diff["only_in_a"], diff.get("name_field_a")),
        ("Только в B", diff["only_in_b"], diff.get("name_field_b")),
    ):
        if not items:
            continue
        sh = wb.create_sheet(title)
        fields = _ordered_fields(items, name_field)
        sh.append(["Товар" if f == name_field else f for f in fields])
        for it in items:
            sh.append([it.get(f) for f in fields])
        _finish_sheet(sh, [50] + [16] * (len(fields) - 1))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def compare_stats_line(diff: dict[str, Any]) -> str:
    """Compact deterministic stats string fed to the LLM for its short answer."""
    mc = diff["match_counts"]
    parts = [
        f"Строк A={diff['count_a']}, B={diff['count_b']}",
        f"совпало {len(diff['matched'])} (точно {mc['exact']}, похоже {mc['fuzzy']}, "
        f"по смыслу {mc['llm']})",
        f"только в A={len(diff['only_in_a'])}, только в B={len(diff['only_in_b'])}",
        f"расхождений по полям={len(diff['mismatches'])}",
    ]
    if diff.get("llm_skipped"):
        parts.append("сопоставление по смыслу пропущено (слишком много строк)")
    if diff["sum_a"] is not None or diff["sum_b"] is not None:
        parts.append(f"сумма A={diff['sum_a']}, сумма B={diff['sum_b']}, разница={diff['delta']}")
    examples = diff["mismatches"][:3]
    if examples:
        ex = "; ".join(
            f"{m.get('name') or m['key_a']} — {m['field']}: {m['value_a']} → {m['value_b']}"
            for m in examples
        )
        parts.append(f"примеры: {ex}")
    return " | ".join(parts)
