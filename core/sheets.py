"""Spreadsheet read/edit helpers over openpyxl, with a multi-format loader.

Deliberately small on the edit side: list sheets, read a range, write a cell or
row, append a row, save. On the LOAD side it sniffs the input format (by magic
bytes first, extension second — Telegram filenames lie) and normalizes every
supported format into one openpyxl ``Workbook`` so all downstream code stays
xlsx-shaped. Supported inputs: ``.xlsx``, ``.xlsm`` (zip), ``.xls`` (legacy
OLE2/BIFF via xlrd), ``.csv`` / ``.tsv`` (utf-8 / utf-8-sig / cp1251). ``.ods``
is intentionally NOT supported (odfpy parsing is non-trivial; skipped by design).
"""

from __future__ import annotations

import csv
import io
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .errors import SheetError

# Content magic bytes — trusted over the (spoofable) filename extension.
_ZIP_MAGIC = b"PK\x03\x04"  # xlsx / xlsm / ods (all are zip containers)
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # legacy .xls (BIFF)

# Delimited-text extensions (parsed only when the extension says so — arbitrary
# texty bytes are NOT silently accepted as CSV, so a corrupt binary still fails).
_TEXT_EXTS = {".csv", ".tsv"}

# Shown to the user (via the excel bot) when a file can't be read.
SUPPORTED_INPUT_FORMATS = (".xlsx", ".xlsm", ".xls", ".csv", ".tsv")


def load(path: str | Path, *, create_if_missing: bool = False) -> Workbook:
    """Load a spreadsheet of any supported format, normalized to a Workbook.

    Sniffs the format from the file's content + extension (see
    :func:`read_bytes_to_workbook`). Optionally create a fresh workbook if absent.

    Raises:
        SheetError: if the file is missing and ``create_if_missing`` is False, or
            if the content is an unsupported/corrupt format.
    """
    p = Path(path)
    if not p.exists():
        if create_if_missing:
            return Workbook()
        raise SheetError(f"workbook not found: {p}")
    return read_bytes_to_workbook(p.read_bytes(), filename=p.name)


def read_bytes_to_workbook(data: bytes, *, filename: str = "") -> Workbook:
    """Sniff raw uploaded bytes and return a normalized openpyxl ``Workbook``.

    Format detection is content-first (magic bytes), extension-second: a zip
    header → xlsx/xlsm, an OLE2 header → legacy .xls, otherwise a ``.csv``/``.tsv``
    extension → delimited text. Anything else (unknown binary, corrupt file,
    ``.ods``) raises :class:`SheetError`.

    Args:
        data: The raw file bytes (e.g. a Telegram document download).
        filename: The original filename — used only as an extension hint for the
            text branch and for the error message.

    Raises:
        SheetError: on an empty, unsupported, or corrupt file.
    """
    if not data:
        raise SheetError(_unsupported(filename))
    ext = Path(filename).suffix.lower()

    if data[:4] == _ZIP_MAGIC:
        return _load_zip(data, filename)
    if data[:8] == _OLE2_MAGIC:
        return _load_xls(data, filename)
    if ext in _TEXT_EXTS:
        return _load_delimited(data, ext, filename)
    raise SheetError(_unsupported(filename))


def _unsupported(filename: str) -> str:
    name = filename or "file"
    return f"unsupported or corrupt spreadsheet: {name} (supported: {', '.join(SUPPORTED_INPUT_FORMATS)})"


def _load_zip(data: bytes, filename: str) -> Workbook:
    """Parse zip-container bytes as .xlsx/.xlsm via openpyxl (rejects .ods)."""
    try:
        names = set(zipfile.ZipFile(io.BytesIO(data)).namelist())
    except zipfile.BadZipFile as exc:
        raise SheetError(_unsupported(filename)) from exc
    if "mimetype" in names or "content.xml" in names:  # ODF spreadsheet, not supported
        raise SheetError(_unsupported(filename))
    try:
        return load_workbook(io.BytesIO(data))
    except Exception as exc:  # BadZipFile / InvalidFileException / KeyError on a broken zip
        raise SheetError(_unsupported(filename)) from exc


def _xls_text_stats(book: Any) -> tuple[int, int]:
    """Count Cyrillic vs high-range-Latin chars over the first sheet's top rows.

    ``lat_hi`` counts U+00C0–U+00FF — the characters a cp1251 Cyrillic byte turns
    into when xlrd misreads it as iso-8859-1 (the mojibake signature). ``cyr``
    counts real Cyrillic (U+0400–U+04FF). Used to decide/verify a cp1251 reopen.
    """
    cyr = lat_hi = 0
    if getattr(book, "nsheets", 0):
        sh = book.sheet_by_index(0)
        for r in range(min(40, sh.nrows)):
            for c in range(sh.ncols):
                v = sh.cell_value(r, c)
                if not isinstance(v, str):
                    continue
                for ch in v:
                    o = ord(ch)
                    if 0x0400 <= o <= 0x04FF:
                        cyr += 1
                    elif 0x00C0 <= o <= 0x00FF:
                        lat_hi += 1
    return cyr, lat_hi


def _needs_cp1251(codepage: Any, cyr: int, lat_hi: int) -> bool:
    """True if the .xls should be reopened as cp1251.

    Triggers when the file has no CODEPAGE record (xlrd then guesses iso-8859-1,
    mangling Russian/Ukrainian text) OR when the decoded text shows the mojibake
    signature: many high-range-Latin chars and essentially no real Cyrillic.
    """
    if not codepage:
        return True
    return lat_hi > 8 and cyr == 0


def _load_xls(data: bytes, filename: str) -> Workbook:
    """Parse legacy .xls (BIFF) bytes via xlrd, copied into an openpyxl Workbook.

    Legacy exports from 1C and similar often omit the CODEPAGE record, so xlrd
    falls back to iso-8859-1 and every Cyrillic name becomes mojibake
    ("Поставщик" -> "Ïîñòàâùèê"). When that is detected the workbook is reopened
    with ``encoding_override="cp1251"`` — but only kept if it actually yields more
    Cyrillic, so a genuine Western file is never made worse.
    """
    import xlrd

    quiet = io.StringIO()  # swallow xlrd's "No CODEPAGE record" stderr chatter
    try:
        book = xlrd.open_workbook(file_contents=data, logfile=quiet)
    except Exception as exc:  # xlrd.XLRDError and friends
        raise SheetError(_unsupported(filename)) from exc
    cyr, lat_hi = _xls_text_stats(book)
    if _needs_cp1251(book.codepage, cyr, lat_hi):
        try:
            fixed = xlrd.open_workbook(file_contents=data, encoding_override="cp1251", logfile=quiet)
        except Exception:  # a bad override must not sink a readable book
            fixed = None
        if fixed is not None and _xls_text_stats(fixed)[0] > cyr:
            book = fixed
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in book.sheets():
        ws = wb.create_sheet(title=(sheet.name or "Sheet")[:31])
        for r in range(sheet.nrows):
            ws.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    if not wb.sheetnames:  # an empty .xls still yields a usable workbook
        wb.create_sheet("Sheet")
    return wb


def _decode_text(data: bytes) -> str:
    """Decode CSV/TSV bytes trying utf-8-sig, utf-8, then cp1251 (Russian legacy)."""
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise SheetError("could not decode text file (tried utf-8, cp1251)")


def _load_delimited(data: bytes, ext: str, filename: str) -> Workbook:
    """Parse .csv/.tsv bytes into a Workbook (delimiter from extension, else sniffed)."""
    text = _decode_text(data)
    if ext == ".tsv":
        delimiter = "\t"
    else:
        try:
            delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
    wb = Workbook()
    ws = wb.active
    for row in csv.reader(io.StringIO(text), delimiter=delimiter):
        ws.append(row)
    return wb


def sheet_names(wb: Workbook) -> list[str]:
    """Return the workbook's sheet names in order."""
    return list(wb.sheetnames)


def _sheet(wb: Workbook, sheet: str | None) -> Worksheet:
    if sheet is None:
        return wb.active
    if sheet not in wb.sheetnames:
        raise SheetError(f"no such sheet: {sheet}")
    return wb[sheet]


def read_range(wb: Workbook, cell_range: str, *, sheet: str | None = None) -> list[list[Any]]:
    """Read a range (e.g. ``"A1:C10"`` or a single cell ``"A1"``) as rows of values.

    openpyxl returns a bare ``Cell`` for a single-cell address, a tuple of
    cells for a single row/column, and a tuple of row-tuples for a 2-D range.
    Normalize all three to ``list[list[Any]]``.
    """
    ws = _sheet(wb, sheet)
    selected = ws[cell_range]
    if not isinstance(selected, tuple):  # single cell
        return [[selected.value]]
    if selected and not isinstance(selected[0], tuple):  # single row or column
        return [[cell.value for cell in selected]]
    return [[cell.value for cell in row] for row in selected]


def write_cell(wb: Workbook, cell: str, value: Any, *, sheet: str | None = None) -> None:
    """Write a single cell (e.g. ``write_cell(wb, "B2", 42)``)."""
    _sheet(wb, sheet)[cell] = value


def write_row(
    wb: Workbook, row_index: int, values: list[Any], *, start_col: int = 1, sheet: str | None = None
) -> None:
    """Overwrite a row's cells starting at ``start_col`` (1-based)."""
    ws = _sheet(wb, sheet)
    for offset, value in enumerate(values):
        ws.cell(row=row_index, column=start_col + offset, value=value)


def _last_data_row(ws: Worksheet) -> int:
    """Return the last row index that holds any non-empty cell (0 if all empty).

    ``ws.max_row`` is unreliable after a range read: ``ws["A1:C5"]`` materializes
    phantom cells and inflates ``max_row``, so ``ws.append`` would leave a gap.
    Scan upward from ``max_row`` for the first row with real data.
    """
    for row_index in range(ws.max_row, 0, -1):
        if any(cell.value is not None for cell in ws[row_index]):
            return row_index
    return 0


def append_row(wb: Workbook, values: list[Any], *, sheet: str | None = None) -> None:
    """Append a row immediately after the last non-empty data row.

    Robust against an inflated ``ws.max_row`` left by a prior range read — the
    new row always lands directly below the real data with no blank gap.
    """
    ws = _sheet(wb, sheet)
    write_row(wb, _last_data_row(ws) + 1, values, sheet=sheet)


def save(wb: Workbook, path: str | Path) -> None:
    """Save the workbook to ``path`` (creates parent dirs)."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
