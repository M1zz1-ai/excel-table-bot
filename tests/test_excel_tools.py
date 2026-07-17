"""Unit tests for excel.tools — sheet-op tools over a REAL temp workbook.

No network: the tools wrap openpyxl directly. These pin the capability contract
that a future unified agent registers as tools.
"""

from __future__ import annotations

from typing import Any

import pytest
from openpyxl import Workbook

from core.errors import SheetError
from excel import tools


def _sample_wb() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["name", "qty", "status"])
    ws.append(["Apple", 10, "active"])
    ws.append(["Banana", 5, "inactive"])
    return wb


def _tools_for(wb: Workbook, changes: list[int] | None = None) -> dict[str, Any]:
    sink = changes if changes is not None else []
    fns = tools.build_sheet_tools(lambda: wb, on_change=lambda: sink.append(1))
    return {fn.__name__: fn for fn in fns}


def test_list_sheets() -> None:
    wb = _sample_wb()
    fns = _tools_for(wb)
    assert fns["list_sheets"]() == ["Sales"]


def test_read_range_round_trip() -> None:
    wb = _sample_wb()
    fns = _tools_for(wb)
    assert fns["read_range"]("A1:C2") == [["name", "qty", "status"], ["Apple", 10, "active"]]


def test_read_single_cell() -> None:
    wb = _sample_wb()
    fns = _tools_for(wb)
    assert fns["read_range"]("B2") == [[10]]


def test_write_cell_mutates_and_persists() -> None:
    wb = _sample_wb()
    changes: list[int] = []
    fns = _tools_for(wb, changes)
    msg = fns["write_cell"]("B2", "42")
    assert "B2" in msg
    assert fns["read_range"]("B2") == [["42"]]
    assert changes, "write_cell must trigger on_change persistence"


def test_append_row_parses_json_array() -> None:
    wb = _sample_wb()
    changes: list[int] = []
    fns = _tools_for(wb, changes)
    fns["append_row"]('["Cherry", 0, "active"]')
    assert fns["read_range"]("A4:C4") == [["Cherry", 0, "active"]]
    assert changes


def test_append_row_accepts_list_and_json_string_equally() -> None:
    """The model may pass a real list OR a JSON-array string — both append the same row."""
    from_string = _sample_wb()
    fns_str = _tools_for(from_string)
    fns_str["append_row"]('["Cherry", 0, "active"]')

    from_list = _sample_wb()
    fns_list = _tools_for(from_list)
    fns_list["append_row"](["Cherry", 0, "active"])

    expected = [["Cherry", 0, "active"]]
    assert fns_str["read_range"]("A4:C4") == expected
    assert fns_list["read_range"]("A4:C4") == expected


def test_write_row_overwrites() -> None:
    wb = _sample_wb()
    fns = _tools_for(wb)
    fns["write_row"](2, '["Apricot", 99, "active"]')
    assert fns["read_range"]("A2:C2") == [["Apricot", 99, "active"]]


def test_append_row_rejects_non_json() -> None:
    wb = _sample_wb()
    fns = _tools_for(wb)
    with pytest.raises(SheetError):
        fns["append_row"]("not json")


def test_append_row_rejects_scalar() -> None:
    """A bare scalar is genuinely bad input and must still be rejected."""
    wb = _sample_wb()
    fns = _tools_for(wb)
    with pytest.raises(SheetError):
        fns["append_row"]("42")


def test_read_unknown_sheet_raises() -> None:
    wb = _sample_wb()
    fns = _tools_for(wb)
    with pytest.raises(SheetError):
        fns["read_range"]("A1", "Ghost")


def test_register_tools_exposes_all_capabilities() -> None:
    """All sheet ops must register on an Agent so a unified agent can route them."""

    class _RecordingAgent:
        def __init__(self) -> None:
            self.registered: list[str] = []

        def tool(self, fn: Any = None, **kw: Any) -> Any:
            self.registered.append(kw.get("name") or getattr(fn, "__name__", ""))
            return fn

    agent = _RecordingAgent()
    tools.register_tools(agent, _sample_wb)
    for expected in ("list_sheets", "read_range", "write_cell", "append_row", "write_row"):
        assert expected in agent.registered


# ---- numeric coercion (RU text-formatted numbers) -----------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("12,50", 12.5),
        ("1 234,56", 1234.56),
        ("1\xa0234,56", 1234.56),  # no-break-space thousands
        ("828,62 EUR", 828.62),
        ("€1 000", 1000.0),
        ("1.234,56", 1234.56),  # EU grouping
        ("1,234.56", 1234.56),  # US grouping
        ("530.72", 530.72),
        ("1.234.567", 1234567.0),  # EU dot thousands
        (42, 42.0),
        (3.5, 3.5),
        ("", None),
        ("  ", None),
        ("abc", None),
        (None, None),
        (True, None),  # a checkbox is not a quantity
    ],
)
def test_coerce_number(raw: Any, expected: Any) -> None:
    assert tools._coerce_number(raw) == expected


# ---- column_stats: deterministic arithmetic over the column -------------


def test_column_stats_sums_text_formatted_numbers() -> None:
    # A column a naive float() would undercount (comma decimals, nbsp, currency)
    # but coercion sums to the file's stated total — the real fix for 530.72 vs 828.62.
    wb = Workbook()
    ws = wb.active
    ws.append(["Сумма Прих"])
    for v in ["100,00", "200,50 €", "1\xa0234,00", "230,22"]:
        ws.append([v])
    ws.append(["Итого:"])  # label row (F row), non-numeric -> skipped
    fns = _tools_for(wb)
    stats = fns["column_stats"]("A2:A6")
    assert stats["sum"] == 1764.72
    assert stats["count"] == 4
    assert stats["skipped"] == 1  # the "Итого:" label, not counted


def test_column_stats_data_range_matches_when_total_row_excluded() -> None:
    # Discrepancy scenario: stated total row vs recompute over data rows only.
    wb = Workbook()
    ws = wb.active
    for v in ["100,00", "200,50", "230,22"]:
        ws.append([v])
    ws.append(["828,62"])  # a total row that does NOT match the data sum
    fns = _tools_for(wb)
    assert fns["column_stats"]("A1:A3")["sum"] == 530.72  # data rows
    assert tools._coerce_number("828,62") == 828.62  # stated total, coerced


def test_column_stats_no_numbers() -> None:
    wb = Workbook()
    wb.active.append(["name", "status"])
    stats = _tools_for(wb)["column_stats"]("A1:B1")
    assert stats == {"count": 0, "sum": 0.0, "skipped": 2,
                     "min": None, "max": None, "mean": None}


def test_column_stats_registered_as_tool() -> None:
    wb = _sample_wb()
    assert "column_stats" in _tools_for(wb)


def test_agent_system_has_total_row_and_column_stats_rules() -> None:
    sys = tools.AGENT_SYSTEM
    assert "column_stats" in sys
    assert "Итого" in sys and "расходится" in sys  # discrepancy rule present
