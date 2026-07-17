"""Ingestion-robustness tests: legacy .xls codepage repair, sparse-header
alignment, and LLM-planned table extraction (model mocked).

No network and no real customer data — grids are synthetic and mirror the shapes
seen in production (1C-style print/accounting exports)."""

from __future__ import annotations

import io

import xlrd
from openpyxl import Workbook, load_workbook

from core import sheets
from excel import engine


def _xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _wb(rows: list[list]) -> Workbook:
    return load_workbook(io.BytesIO(_xlsx(rows)))


# ---- 1. legacy .xls codepage repair ------------------------------------


def test_needs_cp1251_no_codepage():
    assert sheets._needs_cp1251(None, 0, 0) is True  # missing CODEPAGE record
    assert sheets._needs_cp1251(0, 5, 0) is True


def test_needs_cp1251_mojibake_signature_despite_codepage():
    # High-range latin (lat_hi), zero real Cyrillic -> cp1251 misread as latin1.
    assert sheets._needs_cp1251(1252, 0, 30) is True


def test_needs_cp1251_leaves_good_cyrillic_alone():
    assert sheets._needs_cp1251(1251, 40, 0) is False
    assert sheets._needs_cp1251(1200, 40, 0) is False


class _FakeSheet:
    def __init__(self, name: str, grid: list[list]) -> None:
        self.name = name
        self._g = grid
        self.nrows = len(grid)
        self.ncols = max((len(r) for r in grid), default=0)

    def cell_value(self, r: int, c: int):
        row = self._g[r]
        return row[c] if c < len(row) else ""


class _FakeBook:
    def __init__(self, codepage, grid: list[list]) -> None:
        self.codepage = codepage
        self.nsheets = 1
        self._sheets = [_FakeSheet("Sheet1", grid)]

    def sheets(self):
        return self._sheets

    def sheet_by_index(self, i: int):
        return self._sheets[i]


def test_load_xls_reopens_cp1251_on_missing_codepage(monkeypatch):
    mojibake = [["Ïîñòàâùèê", "Íàèìåíîâàíèå"], ["Àïïëå", 10]]  # cp1251 read as latin1
    clean = [["Поставщик", "Наименование"], ["Аппле", 10]]

    def fake_open(file_contents=None, encoding_override=None, logfile=None):
        if encoding_override == "cp1251":
            return _FakeBook(1251, clean)
        return _FakeBook(None, mojibake)  # no CODEPAGE record

    monkeypatch.setattr(xlrd, "open_workbook", fake_open)
    wb = sheets._load_xls(b"\xd0\xcf\x11\xe0fake", "legacy.xls")
    ws = wb.active
    assert ws["A1"].value == "Поставщик"
    assert ws["B1"].value == "Наименование"


def test_load_xls_does_not_downgrade_western_file(monkeypatch):
    western = [["Supplier", "Product"], ["Apple", 10]]

    def fake_open(file_contents=None, encoding_override=None, logfile=None):
        # cp1251 reopen yields no more Cyrillic than the original -> keep original.
        return _FakeBook(None, western)

    monkeypatch.setattr(xlrd, "open_workbook", fake_open)
    wb = sheets._load_xls(b"\xd0\xcf\x11\xe0fake", "w.xls")
    assert wb.active["A1"].value == "Supplier"


# ---- 2. sparse-header positional alignment -----------------------------


def test_named_rows_aligns_under_sparse_header():
    # A gap at column 1 in the header must NOT shift data under the wrong column.
    rows = [
        ["№", "", "Код", "Наименование товара"],
        ["1", " ", "SKU-900", "Пример товара"],
    ]
    out = engine.named_rows(rows, 0, engine.extract_headers(rows, 0))
    assert out == [{"№": "1", "Код": "SKU-900", "Наименование товара": "Пример товара"}]


def test_extract_headers_preserves_order_skips_gaps():
    rows = [["№", "", "Код", "", "Товар"]]
    assert engine.extract_headers(rows, 0) == ["№", "Код", "Товар"]


# ---- 3. LLM-planned table extraction -----------------------------------


def _wide(pairs: list[tuple[int, object]], width: int = 75) -> list:
    r: list = [""] * width
    for c, v in pairs:
        r[c] = v
    return r


def _x2_like_grid() -> list[list]:
    """Sparse 2-up print export: header@7, data@9, columns duplicated at cols 40+."""
    first = [(1, "№"), (3, "Артикул"), (6, "Товар"), (22, "Кількість"), (27, "Ціна"), (33, "Сума")]
    dup = [(c + 39, v) for c, v in first]  # horizontal duplicate block
    header = _wide(first + dup)

    def data(n, art, name, qty, price, total):
        block = [(1, n), (3, art), (6, name), (22, qty), (27, price), (33, total)]
        return _wide(block + [(c + 39, v) for c, v in block])

    return [
        _wide([(1, "Видаткова накладна № 0001")]),   # 0 metadata
        _wide([(1, "Постачальник:")]),                  # 1
        _wide([]),                                       # 2 blank
        _wide([]),                                       # 3
        _wide([]),                                       # 4
        _wide([]),                                       # 5
        _wide([(1, "Адреса доставки:")]),               # 6
        header,                                          # 7 header
        _wide([]),                                       # 8 blank
        data("1", "SKU-001", "Widget Wax", 2.0, 1.04, 2.08),   # 9 data1
        data("1", "SKU-001", "Widget Wax", 2.0, 1.04, 2.08),   # 10 duplicate -> deduped
        data("2", "SKU-002", "Widget Shampoo", 24.0, 0.97, 23.28),  # 11 data2
        _wide([(1, "№"), (3, "Артикул"), (6, "Товар"), (22, "Кількість"), (27, "Ціна"), (33, "Сума")]),  # 12 repeated header
        data("", "", "Итого", "", "", 25.36),           # 13 total row
    ]


def _x2_plan() -> dict:
    return {
        "header_row": 7,
        "data_start_row": 9,
        "columns": [
            {"col": 1, "field": "№"},
            {"col": 3, "field": "Артикул"},
            {"col": 6, "field": "Товар"},
            {"col": 22, "field": "Кількість"},
            {"col": 27, "field": "Ціна"},
            {"col": 33, "field": "Сума"},
        ],
        "notes": "first block only",
    }


def test_extract_by_plan_data_end_row_trims_footer():
    # A UA totals + payment footer AFTER the items must be excluded by data_end_row.
    grid = _x2_like_grid()
    grid.append(_wide([(1, "Всього найменувань 2, на суму 25,36 Є.")]))  # row 14
    grid.append(_wide([(1, "Увага на наступній сторінці акт звірки!!!!")]))  # row 15
    plan = _x2_plan()
    plan["data_end_row"] = 11  # last real item row
    headers, data = engine.extract_by_plan(grid, plan)
    assert len(data) == 2
    assert all("Всього" not in str(r.get("№", "")) for r in data)


def test_extract_by_plan_first_block_dedupe_total_repeated_header():
    headers, data = engine.extract_by_plan(_x2_like_grid(), _x2_plan())
    assert headers == ["№", "Артикул", "Товар", "Кількість", "Ціна", "Сума"]
    # duplicate row 10, repeated header 12, and total 13 all dropped -> 2 items.
    assert len(data) == 2
    assert data[0]["Артикул"] == "SKU-001" and data[1]["Артикул"] == "SKU-002"
    # only the FIRST block columns are read (no cols from the duplicated block).
    assert set(data[0].keys()) == {"№", "Артикул", "Товар", "Кількість", "Ціна", "Сума"}
    assert data[1]["Сума"] == 23.28


def test_table_plan_prompt_preserves_column_indices():
    rows = _x2_like_grid()
    prompt = engine.table_plan_prompt(rows)
    # The Kількість caption sits at column 22 — the prompt must expose that index.
    assert '"22"' in prompt and "Кількість" in prompt
    assert "Видаткова накладна № 0001" in prompt  # metadata row 0 present


def test_load_table_planned_uses_plan_when_available():
    wb = _wb([["ignored heuristic header"], ["x"], ["name", "qty"], ["Apple", 10], ["Banana", 5]])

    def plan_fn(rows):
        return {
            "header_row": 2,
            "data_start_row": 3,
            "columns": [{"col": 0, "field": "name"}, {"col": 1, "field": "qty"}],
            "notes": "",
        }

    headers, data = engine.load_table_planned(wb, plan_fn=plan_fn)
    assert headers == ["name", "qty"]
    assert [r["name"] for r in data] == ["Apple", "Banana"]


def test_load_table_planned_falls_back_when_plan_raises():
    wb = _wb([["name", "qty"], ["Apple", 10]])

    def boom(rows):
        raise RuntimeError("model down")

    headers, data = engine.load_table_planned(wb, plan_fn=boom)
    assert headers == ["name", "qty"] and data[0]["name"] == "Apple"


def test_load_table_planned_falls_back_when_plan_empty():
    wb = _wb([["name", "qty"], ["Apple", 10]])

    def empty(rows):
        return {"header_row": 0, "data_start_row": 99, "columns": [], "notes": ""}

    headers, data = engine.load_table_planned(wb, plan_fn=empty)
    assert headers == ["name", "qty"] and data[0]["name"] == "Apple"


def test_load_table_planned_no_plan_fn_uses_heuristic():
    wb = _wb([["name", "qty"], ["Apple", 10]])
    headers, data = engine.load_table_planned(wb)
    assert headers == ["name", "qty"] and len(data) == 1
