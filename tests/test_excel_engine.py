"""Deterministic reformat/compare engine tests (no LLM, no network)."""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from excel import engine


def _xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _wb(rows: list[list]) -> Workbook:
    return load_workbook(io.BytesIO(_xlsx(rows)))


# ---- tabular parsing ----------------------------------------------------


def test_detect_header_row_skips_metadata():
    rows = [["Отчёт за май", None, None], ["", "", ""], ["name", "qty", "status"], ["Apple", 10, "x"]]
    assert engine.detect_header_row(rows) == 2


def test_load_table_drops_total_row():
    wb = _wb([["name", "qty"], ["Apple", 10], ["Banana", 5], ["Итого:", 15]])
    headers, data = engine.load_table(wb)
    assert headers == ["name", "qty"]
    assert [r["name"] for r in data] == ["Apple", "Banana"]  # Итого row excluded


def test_is_total_row():
    assert engine.is_total_row({"a": "Итого:", "b": 15})
    assert engine.is_total_row({"a": "Total", "b": 1})
    assert not engine.is_total_row({"a": "Apple", "b": 10})


# ---- reformat -----------------------------------------------------------


def test_parse_template_mandatory_and_cid():
    wb = _wb([["*Name*", "Qty", "CID"], ["", "", "C-100"]])
    tpl = engine.parse_template(wb)
    assert tpl["headers"] == [
        {"name": "Name", "mandatory": True},
        {"name": "Qty", "mandatory": False},
        {"name": "CID", "mandatory": False},
    ]
    assert tpl["cid_constant"] == "C-100"


def test_apply_mapping_template_cols_only_with_cid_constant():
    tpl = {
        "headers": [
            {"name": "Name", "mandatory": True},
            {"name": "Qty", "mandatory": False},
            {"name": "CID", "mandatory": False},
        ],
        "cid_constant": "C-100",
    }
    source = [{"Товар": "Apple", "Кол": 10}, {"Товар": "Banana", "Кол": 5}]
    plan = {
        "columns": [
            {"template_col": "Name", "source_col": "Товар", "constant": "", "fuzzy": True},
            {"template_col": "Qty", "source_col": "Кол", "constant": "", "fuzzy": False},
        ],
        "key_column": "Name",
    }
    rows, notes = engine.apply_mapping(tpl, source, plan)
    assert rows[0] == {"Name": "Apple", "Qty": 10, "CID": "C-100"}  # CID from template
    assert notes["fuzzy_fields"] == ["Name"]
    assert notes["duplicate_keys"] == 0
    assert notes["row_count"] == 2


def test_parse_template_planned_skips_metadata_rows():
    # A real template: title row + blank row above the actual *mandatory*/CID header.
    wb = _wb([
        ["Шаблон загрузки", None, None],
        [None, None, None],
        ["*Артикул*", "Название", "CID"],
        ["", "", "SHOP-7"],
    ])

    def plan_fn(rows):
        return {"header_row": 2, "data_start_row": 3,
                "columns": [{"col": 0, "field": "Артикул"}], "notes": ""}

    tpl = engine.parse_template_planned(wb, plan_fn=plan_fn)
    assert tpl["headers"] == [
        {"name": "Артикул", "mandatory": True},
        {"name": "Название", "mandatory": False},
        {"name": "CID", "mandatory": False},
    ]
    assert tpl["cid_constant"] == "SHOP-7"


def test_parse_template_planned_falls_back_on_plan_failure():
    wb = _wb([["*Name*", "Qty"], ["", ""]])

    def boom(rows):
        raise RuntimeError("model down")

    tpl = engine.parse_template_planned(wb, plan_fn=boom)  # falls back to row 0
    assert tpl["headers"][0] == {"name": "Name", "mandatory": True}


def test_apply_mapping_fill_stats_and_quality():
    tpl = {
        "headers": [
            {"name": "Name", "mandatory": True},
            {"name": "Qty", "mandatory": False},
            {"name": "Price", "mandatory": False},
        ],
        "cid_constant": None,
    }
    source = [{"Товар": "Apple", "Кол": 10}, {"Товар": "Banana", "Кол": 5}]
    plan = {
        "columns": [
            {"template_col": "Name", "source_col": "Товар", "constant": "", "fuzzy": False},
            {"template_col": "Qty", "source_col": "Кол", "constant": "", "fuzzy": False},
        ],
        "key_column": "Name",
    }
    rows, notes = engine.apply_mapping(tpl, source, plan)
    assert notes["template_cols"] == 3 and notes["mapped_columns"] == 2
    assert notes["unmapped_columns"] == ["Price"]
    assert notes["empty_columns"] == ["Price"]  # Price never filled
    assert notes["column_fill"] == {"Name": 2, "Qty": 2, "Price": 0}
    assert notes["overall_fill_pct"] == round(4 / 6, 4)  # 4 of 6 cells filled
    assert engine.reformat_quality(notes) == "ok"  # 1 of 3 unmapped = 33% <= 50%


def test_reformat_quality_empty_and_weak():
    # nothing mapped -> empty
    empty = {"row_count": 2, "mapped_columns": 0, "overall_fill_pct": 0.0,
             "template_cols": 3, "unmapped_columns": ["A", "B", "C"]}
    assert engine.reformat_quality(empty) == "empty"
    # >50% of template columns unmapped -> weak
    weak = {"row_count": 2, "mapped_columns": 1, "overall_fill_pct": 0.33,
            "template_cols": 3, "unmapped_columns": ["B", "C"]}
    assert engine.reformat_quality(weak) == "weak"
    # zero rows -> empty
    assert engine.reformat_quality({"row_count": 0, "mapped_columns": 2,
                                    "overall_fill_pct": 0.0, "template_cols": 2,
                                    "unmapped_columns": []}) == "empty"


def test_apply_mapping_tolerates_decorated_template_col():
    # Model echoes "Name (mandatory)" instead of "Name"; must still map (was a silent
    # empty-column bug for mandatory columns).
    tpl = {"headers": [{"name": "Name", "mandatory": True}], "cid_constant": None}
    plan = {"columns": [
        {"template_col": "Name (mandatory)", "source_col": "Товар", "constant": "", "fuzzy": False}
    ], "key_column": ""}
    rows, notes = engine.apply_mapping(tpl, [{"Товар": "Apple"}], plan)
    assert rows[0] == {"Name": "Apple"}  # filled, not None
    assert notes["unmapped_columns"] == [] and notes["mapped_columns"] == 1


def test_apply_mapping_flags_unmapped_mandatory():
    tpl = {"headers": [{"name": "Price", "mandatory": True}], "cid_constant": None}
    plan = {"columns": [], "key_column": ""}
    _, notes = engine.apply_mapping(tpl, [{"x": 1}], plan)
    assert notes["unmapped_mandatory"] == ["Price"]


def test_build_reformat_xlsx_roundtrip():
    tpl = {"headers": [{"name": "Name", "mandatory": False}, {"name": "Qty", "mandatory": False}]}
    data = engine.build_reformat_xlsx(tpl, [{"Name": "Apple", "Qty": 10}])
    ws = load_workbook(io.BytesIO(data)).active
    assert [c.value for c in ws[1]] == ["Name", "Qty"]
    assert [c.value for c in ws[2]] == ["Apple", 10]


def test_reformat_end_to_end_realistic_fixture():
    """A realistic messy source: title/blank metadata rows, a total row, RU headers,
    reshaped into a *Name*/Price/CID template — the flow Reformat mode has to survive."""
    # Template: mandatory Name, Price, and a CID constant column.
    tpl_wb = _wb([["*Товар*", "Цена", "CID"], ["", "", "SHOP-7"]])
    template = engine.parse_template(tpl_wb)

    # Source has 2 metadata rows above the header and an Итого total row below.
    src_wb = _wb(
        [
            ["Прайс-лист магазина", None, None],
            [None, None, None],
            ["Наименование", "Стоимость", "Остаток"],
            ["Молоток", "150,50", 12],
            ["Гвозди 100мм", "1 200,00", 300],
            ["Итого:", "1 350,50", 312],
        ]
    )
    headers, rows = engine.load_table(src_wb)
    assert headers == ["Наименование", "Стоимость", "Остаток"]
    assert len(rows) == 2  # Итого row dropped

    plan = {
        "columns": [
            {"template_col": "Товар", "source_col": "Наименование", "constant": "", "fuzzy": False},
            {"template_col": "Цена", "source_col": "Стоимость", "constant": "", "fuzzy": False},
        ],
        "key_column": "Товар",
    }
    out_rows, notes = engine.apply_mapping(template, rows, plan)
    assert notes["row_count"] == 2
    assert notes["unmapped_mandatory"] == []  # Товар is mapped
    assert notes["duplicate_keys"] == 0

    data = engine.build_reformat_xlsx(template, out_rows)
    ws = load_workbook(io.BytesIO(data)).active
    assert [c.value for c in ws[1]] == ["Товар", "Цена", "CID"]
    assert [c.value for c in ws[2]] == ["Молоток", "150,50", "SHOP-7"]  # CID constant applied
    assert [c.value for c in ws[3]] == ["Гвозди 100мм", "1 200,00", "SHOP-7"]


# ---- compare ------------------------------------------------------------


def _compare_plan() -> dict:
    return {
        "key_a": "id",
        "key_b": "id",
        "compare_columns": [{"a": "amount", "b": "amount", "label": "amount"}],
        "sum_column_a": "amount",
        "sum_column_b": "amount",
    }


def test_diff_tables_join_mismatch_and_sums():
    rows_a = [{"id": "1", "amount": "100"}, {"id": "2", "amount": "200"}]
    rows_b = [{"id": "1", "amount": "100"}, {"id": "2", "amount": "250"}, {"id": "3", "amount": "50"}]
    diff = engine.diff_tables(rows_a, rows_b, _compare_plan())
    assert [r["key"] for r in diff["only_in_b"]] == ["3"]
    assert diff["only_in_a"] == []
    assert len(diff["mismatches"]) == 1
    m = diff["mismatches"][0]
    assert m["key_a"] == "2" and m["key_b"] == "2" and m["delta"] == -50.0
    assert m["match_method"] == "exact"
    assert diff["sum_a"] == 300.0 and diff["sum_b"] == 400.0 and diff["delta"] == -100.0
    assert diff["match_counts"] == {"exact": 2, "fuzzy": 0, "llm": 0}


def test_diff_forces_sum_column_into_compare_set():
    # Plan omits the sum column from compare_columns; the diff must still surface the
    # money difference (label «Сумма») so «Расхождения» never hides it.
    plan = {
        "key_a": "id", "key_b": "id",
        "compare_columns": [],  # LLM forgot to compare the sum
        "sum_column_a": "amount", "sum_column_b": "amount",
    }
    diff = engine.diff_tables(
        [{"id": "1", "amount": "100"}], [{"id": "1", "amount": "150"}], plan
    )
    assert any(m["field"] == "Сумма" and m["delta"] == -50.0 for m in diff["mismatches"])


def test_diff_coerces_ru_text_numbers():
    # "1 234,50" vs 1234.5 must NOT register as a mismatch.
    rows_a = [{"id": "1", "amount": "1 234,50"}]
    rows_b = [{"id": "1", "amount": 1234.5}]
    diff = engine.diff_tables(rows_a, rows_b, _compare_plan())
    assert diff["mismatches"] == []


def test_build_compare_xlsx_расхождения_first_then_почему_then_итог():
    diff = engine.diff_tables(
        [{"id": "1", "name": "Товар1", "amount": "100"}],
        [{"id": "1", "name": "Товар1", "amount": "150"}, {"id": "2", "name": "Товар2", "amount": "5"}],
        _compare_plan(),
    )
    wb = load_workbook(io.BytesIO(engine.build_compare_xlsx(diff)))
    # Per-product diff is what the user opens for -> first sheet, active on open.
    assert wb.sheetnames[:3] == ["Расхождения", "Почему разница", "Итог"]
    assert wb.active.title == "Расхождения"
    assert "Только в B" in wb.sheetnames
    assert "Только в A" not in wb.sheetnames  # empty -> omitted
    assert wb["Расхождения"].max_row == 2  # header + 1 differing field
    assert wb["Итог"].freeze_panes == "A2"  # header frozen
    # «Итог» ends with a pointer to the detail sheets.
    итог_flat = [v for row in wb["Итог"].iter_rows(values_only=True) for v in row]
    assert "Строк в A" in итог_flat and 1 in итог_flat
    assert any(v and "Детали по товарам" in str(v) for v in итог_flat)


def test_build_compare_xlsx_почему_разница_sheet():
    diff = engine.diff_tables(
        [{"id": "1", "name": "Товар1", "amount": "100"}, {"id": "9", "name": "ТолькоA", "amount": "30"}],
        [{"id": "1", "name": "Товар1", "amount": "150"}],
        _compare_plan(),
    )
    wb = load_workbook(io.BytesIO(engine.build_compare_xlsx(diff)))
    assert wb.sheetnames[1] == "Почему разница"  # after «Расхождения»
    sh = wb["Почему разница"]
    assert [c.value for c in sh[1]] == ["Товар", "Артикул", "Вклад в разницу", "Причина"]
    # bold total row equals Δ
    last = [c.value for c in sh[sh.max_row]]
    assert last[0] == "ИТОГО разница" and last[2] == diff["delta"]
    assert sh[sh.max_row][0].font.bold is True


def test_build_compare_xlsx_расхождения_is_product_centric():
    diff = engine.diff_tables(
        [{"id": "SKU-100", "name": "Пример товара", "amount": "17.18"}],
        [{"id": "SKU-100", "name": "Пример товара", "amount": "15.12"}],
        _compare_plan(),
    )
    wb = load_workbook(io.BytesIO(engine.build_compare_xlsx(diff)))
    sh = wb["Расхождения"]
    assert [c.value for c in sh[1]] == ["Товар", "Код", "Поле", "Значение A", "Значение B", "Разница"]
    row = [c.value for c in sh[2]]
    assert row[0] == "Пример товара" and row[1] == "SKU-100"  # product name + code, not a bare key


def test_build_compare_xlsx_совпавшие_sheet_for_nonexact():
    def fake_llm(keys_a, keys_b):
        return [{"a": "гвоздь", "b": "цвях"}]

    diff = engine.diff_tables(
        [{"id": "гвоздь", "name": "Гвоздь строительный", "amount": "1"}],
        [{"id": "цвях", "name": "Цвях будівельний", "amount": "1"}],
        _compare_plan(),
        llm_pair=fake_llm,
    )
    wb = load_workbook(io.BytesIO(engine.build_compare_xlsx(diff)))
    assert "Совпавшие (неточно)" in wb.sheetnames
    sh = wb["Совпавшие (неточно)"]
    assert [c.value for c in sh[1]] == ["Товар", "Код A", "Код B", "Метод", "Уверенность"]
    assert sh.cell(row=2, column=4).value == "по смыслу"  # method rendered in Russian


def test_compare_stats_line_russian_with_sums_and_examples():
    diff = engine.diff_tables(
        [{"id": "1", "name": "Товар1", "amount": "100"}],
        [{"id": "1", "name": "Товар1", "amount": "150"}],
        _compare_plan(),
    )
    line = engine.compare_stats_line(diff)
    assert "сумма A=100" in line and "разница=-50" in line and "примеры:" in line
    assert "совпало 1" in line and "точно 1" in line
    assert "Товар1" in line  # example is product-centric, not a bare key
