"""Wizard state-machine tests: reformat/compare flows through ExcelBot.

The LLM plan/answer agent is faked (bot._plan_agent monkeypatched); files are
real xlsx bytes parsed by core.sheets; the session uses an in-memory state store
and a tmp workdir. No network.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from excel import i18n
from excel.bot import ExcelBot
from excel.session import ExcelSession


def _xlsx(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _FakeState:
    def __init__(self) -> None:
        self.store: dict[str, Any] = {}

    def get_session(self, key: str, default: Any = None) -> Any:
        return self.store.get(key, default)

    def set_session(self, key: str, value: Any, **kw: Any) -> None:
        self.store[key] = value

    def clear_session(self, key: str) -> None:
        self.store.pop(key, None)


class _FakeTg:
    def __init__(self) -> None:
        self.texts: list[str] = []
        self.docs: list[dict[str, Any]] = []

    async def send_text(self, text: str, chat_id: int | None = None, **kw: Any) -> Any:
        self.texts.append(text)

    async def send_document(self, document: Any, chat_id: int | None = None, **kw: Any) -> Any:
        self.docs.append({"bytes": document, "filename": kw.get("filename"), "caption": kw.get("caption")})


class _FakePlanAgent:
    def __init__(self, structured: dict, run_reply: str) -> None:
        self._structured = structured
        self._run = run_reply

    def structured_output(self, prompt: str, schema: dict) -> dict:
        return self._structured

    def run(self, prompt: str) -> str:
        return self._run


def _bot(tmp_path, plan: dict, answer: str = "answer"):
    tg = _FakeTg()
    sess = ExcelSession(_FakeState(), workdir=tmp_path)
    bot = ExcelBot(tg, object(), sess, owner_chat_id=1)  # type: ignore[arg-type]
    bot._plan_agent = lambda system: _FakePlanAgent(plan, answer)  # type: ignore[method-assign]
    return bot, tg, sess


# ---- reformat happy path ------------------------------------------------


REFORMAT_PLAN = {
    "columns": [
        {"template_col": "Name", "source_col": "Товар", "constant": "", "fuzzy": False},
        {"template_col": "Qty", "source_col": "Кол", "constant": "", "fuzzy": False},
    ],
    "key_column": "Name",
}


@pytest.mark.asyncio
async def test_reformat_full_flow(tmp_path):
    bot, tg, sess = _bot(tmp_path, REFORMAT_PLAN)

    await bot.on_text(7, i18n.BTN_REFORMAT)  # start
    assert sess.wizard_get(7)["step"] == 1
    assert "1/3" in tg.texts[-1]

    await bot.on_document(7, _xlsx([["*Name*", "Qty", "CID"], ["", "", "C-100"]]), "tpl.xlsx")
    assert sess.wizard_get(7)["step"] == 2
    assert "2/3" in tg.texts[-1]

    await bot.on_document(7, _xlsx([["Товар", "Кол"], ["Apple", 10], ["Banana", 5]]), "data.xlsx")
    assert sess.wizard_get(7)["step"] == 3
    assert "3/3" in tg.texts[-1]

    await bot.on_text(7, "перенеси данные в шаблон")
    assert sess.wizard_get(7) is None  # cleared after completion
    doc = tg.docs[-1]
    assert doc["filename"].startswith("reformatted_data_") and doc["filename"].endswith(".xlsx")
    ws = load_workbook(io.BytesIO(doc["bytes"])).active
    assert [c.value for c in ws[1]] == ["Name", "Qty", "CID"]
    assert [c.value for c in ws[2]] == ["Apple", 10, "C-100"]  # CID constant applied


@pytest.mark.asyncio
async def test_reformat_empty_mapping_sends_no_file_only_explanation(tmp_path):
    # A mapping plan that maps nothing -> the bot must NOT send a silent empty file.
    bot, tg, sess = _bot(tmp_path, {"columns": [], "key_column": ""})
    await bot.on_text(7, i18n.BTN_REFORMAT)
    await bot.on_document(7, _xlsx([["*Name*", "Qty"], ["", ""]]), "tpl.xlsx")
    await bot.on_document(7, _xlsx([["Товар", "Кол"], ["Apple", 10]]), "data.xlsx")
    await bot.on_text(7, "непонятная инструкция")

    assert tg.docs == []  # no useless file
    msg = tg.texts[-1]
    assert "Не удалось сопоставить" in msg
    assert "Колонки шаблона" in msg and "Name" in msg
    assert "Колонки в вашем файле" in msg and "Товар" in msg


# ---- compare happy path -------------------------------------------------


COMPARE_PLAN = {
    "key_a": "id",
    "key_b": "id",
    "compare_columns": [{"a": "amount", "b": "amount", "label": "amount"}],
    "sum_column_a": "amount",
    "sum_column_b": "amount",
}


@pytest.mark.asyncio
async def test_compare_full_flow_emits_summary_and_xlsx(tmp_path):
    bot, tg, sess = _bot(tmp_path, COMPARE_PLAN, answer="Totals differ by 100.")

    await bot.on_text(7, i18n.BTN_COMPARE)
    await bot.on_document(7, _xlsx([["id", "amount"], [1, 100], [2, 200]]), "a.xlsx")
    await bot.on_document(7, _xlsx([["id", "amount"], [1, 100], [2, 250], [3, 50]]), "b.xlsx")
    await bot.on_text(7, "почему итоговая сумма разная")

    assert sess.wizard_get(7) is None
    doc = tg.docs[-1]
    assert doc["filename"] == "compare_report.xlsx"
    # Only-in-A is empty here (both files share ids 1,2), so its sheet is omitted;
    # «Расхождения» is first (active on open), then «Почему разница», then «Итог».
    assert load_workbook(io.BytesIO(doc["bytes"])).sheetnames == [
        "Расхождения", "Почему разница", "Итог", "Только в B"
    ]
    summary = tg.texts[-1]
    assert "Δ = -100" in summary and "Totals differ by 100." in summary
    assert "совпало 2" in summary  # 2 exact joins mentioned


# ---- wrong input at a step ----------------------------------------------


@pytest.mark.asyncio
async def test_text_at_file_step_nudges_not_advances(tmp_path):
    bot, tg, sess = _bot(tmp_path, REFORMAT_PLAN)
    await bot.on_text(7, i18n.BTN_REFORMAT)  # step 1 expects a file
    await bot.on_text(7, "some text instead of a file")
    assert tg.texts[-1] == i18n.t("ru", "need_file")
    assert sess.wizard_get(7)["step"] == 1  # not advanced


@pytest.mark.asyncio
async def test_file_at_instruction_step_nudges(tmp_path):
    bot, tg, sess = _bot(tmp_path, COMPARE_PLAN)
    await bot.on_text(7, i18n.BTN_COMPARE)
    await bot.on_document(7, _xlsx([["id", "amount"], [1, 1]]), "a.xlsx")
    await bot.on_document(7, _xlsx([["id", "amount"], [1, 1]]), "b.xlsx")  # now step 3
    await bot.on_document(7, _xlsx([["id"], [1]]), "c.xlsx")  # a file at the text step
    assert tg.texts[-1] == i18n.t("ru", "need_text")
    assert sess.wizard_get(7)["step"] == 3


# ---- cancel / reset -----------------------------------------------------


@pytest.mark.asyncio
async def test_cancel_clears_wizard(tmp_path):
    bot, tg, sess = _bot(tmp_path, REFORMAT_PLAN)
    await bot.on_text(7, i18n.BTN_REFORMAT)
    await bot.on_cancel(7)
    assert sess.wizard_get(7) is None
    assert tg.texts[-1] == i18n.t("ru", "cancelled")


@pytest.mark.asyncio
async def test_help_button_resets_active_wizard(tmp_path):
    bot, tg, sess = _bot(tmp_path, REFORMAT_PLAN)
    await bot.on_text(7, i18n.BTN_REFORMAT)
    await bot.on_text(7, i18n.BTN_HELP)  # a menu button mid-wizard
    assert sess.wizard_get(7) is None  # wizard reset
    assert tg.texts[-1] == i18n.t("ru", "help")


@pytest.mark.asyncio
async def test_wizard_temp_files_removed_on_clear(tmp_path):
    bot, tg, sess = _bot(tmp_path, REFORMAT_PLAN)
    await bot.on_text(7, i18n.BTN_REFORMAT)
    await bot.on_document(7, _xlsx([["*Name*", "Qty"], ["", ""]]), "tpl.xlsx")
    saved = Path(sess.wizard_get(7)["files"]["tpl"])
    assert saved.exists()
    await bot.on_cancel(7)
    assert not saved.exists()
